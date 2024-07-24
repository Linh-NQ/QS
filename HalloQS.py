#!/usr/bin/env python
# coding: utf-8

# In[1]:


from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import re
from io import StringIO
output_string = StringIO()

import pandas as pd
import numpy as np
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

#from openpyxl import load_workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import os
import glob


# In[2]:


# Dictionary für Zugangsdaten
login_dict = {
    '999997122': '7122@invenz',
    '999911407': 'q2-sgLcxV?'
}

# Dictionary für Starnet-Login aller Abteilungen bei Rechnungen
login_benutzername = {
    'F&E': '999911405',
    'QS': '999997122',
    'DM': '999911407',
    'P': '999911406'
}

login_passwort = {
    'F&E': 'g7*WyR6TNG',
    'QS': '7122@invenz',
    'DM': 'q2-sgLcxV?',
    'P': 'R4/LqP3ALC'    
}


# In[3]:


# Zurück-Button
def goback():
    root2.grid_forget()
    labor.grid(row=5, column=1)
    rechnungen_check.grid(row=7, column=1)


# Funktion für Labor 28 Einträge
def run_eintrag():
    labor.grid_forget()
    rechnungen_check.grid_forget()
    root_run = Frame(root, width=700, height=200, bg = '#eeeee4')
    root_run.grid(row=2, column=1)

    # Zuürck-Button
    def back_homepage():
        root_run.grid_forget()
        labor.grid(row=5, column=1)
        rechnungen_check.grid(row=7, column=1)
        
    back_homepage_button = Button(root_run, text = 'Zurück', bg = '#869287',
                                  font=('Ink free',10,'bold'), command = back_homepage)
    back_homepage_button.grid(row=30, column=7)
    
    # User Interface Layout
    text_title = Label(root_run, text='Labor 28 Einträge', bg = '#eeeee4', font=('Ink free',14,'bold'))
    text_title.grid(row=1, column=1)
    
    # Dropdown: welche Login-Daten?
    login_text = Label(root_run, text='Wähle den entsprechenden Nutzer für das Login aus', bg = '#eeeee4', font=('Ink free',13))
    login_text.grid(row=2, column=1, columnspan=5)
    options = ['999997122', '999911407']
    selected_option = StringVar()
    dropdown_user = ttk.Combobox(root_run, textvariable=selected_option, values=options, width=8)
    dropdown_user.grid(row=2, column=6)
    
    # 1. Methode: Daten im GUI eingeben
    method_one_text = Label(root_run, text='\n1. Methode: Gib hier die Daten ein',
                            bg = '#eeeee4', font=('Ink free',12,'bold'))
    method_one_text.grid(row=3, column=1, columnspan=5)
    
    # Dropdown: wie viele verschiedene Proben?
    drop_text = Label(root_run, text='Anzahl:', bg = '#eeeee4', font=('Ink free',13))
    drop_text.grid(row=4, column=1)
    options = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
    selected_option = StringVar()
    dropdown = ttk.Combobox(root_run, textvariable=selected_option, values=options, width=2)
    dropdown.grid(row=4, column=2)
    dropdown.current(1)
    
    def on_select(event):
        global number, start_row, widget_probenname_vor, widget_probenname_nach, widget_nummerierung, widget_nummerierung, widget_nummerierung2, widget_parameter 
        number = int(dropdown.get())
        widget_probenname_vor = []
        widget_probenname_nach = []
        widget_nummerierung = []
        widget_nummerierung2 = []
        widget_parameter = []
        start_row = 0
        
        # create scrollbar
        def on_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        frame = Frame(root_run, width=700, height=200, bg = '#eeeee4')
        frame.grid(row=5, column=1, sticky="nsew", columnspan=5)
        canvas = Canvas(frame, bg = '#eeeee4')
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar = Scrollbar(frame, command=canvas.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=scrollbar.set)
        # Bind the canvas to the frame's size
        canvas.bind("<Configure>", on_configure)
        # Create a frame inside the canvas to hold the widgets
        inner_frame = Frame(canvas, bg = '#eeeee4')
        canvas.create_window((0, 0), window=inner_frame, anchor="nw")
            
        
        for i in range(number):
            probe = Label(inner_frame, text='Probenbezeichnung:', bg = '#eeeee4', font=('Ink free',13))
            probe.grid(row=3+start_row, column=1)
            probe_vorname = Label(inner_frame, text='Vorname', bg = '#eeeee4', font=('Ink free',10,'italic'))
            probe_vorname.grid(row=3+start_row, column=2, columnspan=2)
            probe_nachname = Label(inner_frame, text='Nachname', bg = '#eeeee4', font=('Ink free',10,'italic'))
            probe_nachname.grid(row=3+start_row, column=5, columnspan=3)        
            probe_entry_vor = Entry(inner_frame, font=('Arial',10), width=9)
            probe_entry_vor.grid(row=4+start_row, column=2, columnspan=3)
            probe_entry_nach = Entry(inner_frame, font=('Arial',10), width=14)
            probe_entry_nach.grid(row=4+start_row, column=5, columnspan=10)

            nr = Label(inner_frame, text='Nummerierung:', bg = '#eeeee4', font=('Ink free',13))
            nr.grid(row=5+start_row, column=1)
            nr_entry = Entry(inner_frame, font=('Arial',10), width=3)
            nr_entry.grid(row=5+start_row, column=2)
            strich = Label(inner_frame, text='-', bg = '#eeeee4', font=('Ink free',13))
            strich.grid(row=5+start_row, column=3)
            nr_entry2 = Entry(inner_frame, font=('Arial',10), width=3)
            nr_entry2.grid(row=5+start_row, column=4)
            parameter = Label(inner_frame, text='Parameter:', bg = '#eeeee4', font=('Ink free',13))
            parameter.grid(row=6+start_row, column=1)
            parameter_entry = Entry(inner_frame, font=('Arial',10), width=23)
            parameter_entry.grid(row=6+start_row, column=2, columnspan=12)
            space = Label(inner_frame, text='', bg = '#eeeee4')
            space.grid(row=7+start_row, column=1)
            start_row +=5
            
            widget_probenname_vor.append(probe_entry_vor)
            widget_probenname_nach.append(probe_entry_nach)
            widget_nummerierung.append(nr_entry)
            widget_nummerierung2.append(nr_entry2)
            widget_parameter.append(parameter_entry)
                
        # Make the rows and columns expand when the window is resized
        root_run.grid_rowconfigure(0, weight=1)
        root_run.grid_columnconfigure(0, weight=1)
        
        # Go Dodo Button und Funktion
        def go_dodo():
            global widget_probenname_vor, widget_probenname_nach, widget_nummerierung, widget_nummerierung2, widget_parameter
            for i in range(number):
                widget_probenname_vor[i] = widget_probenname_vor[i].get()
                widget_probenname_nach[i] = widget_probenname_nach[i].get()
                widget_parameter[i] = widget_parameter[i].get()
                # widget_parameter[i] kann mehrere Parameter enthalten, die mit ; getrennt sind
                widget_parameter[i] = widget_parameter[i].split(';')

                widget_nummerierung[i] = widget_nummerierung[i].get()
                widget_nummerierung2[i] = widget_nummerierung2[i].get()
                # wenn widget_nummerierung ganzzahlig ist:
                if ('.' not in widget_nummerierung[i]) & ('.' not in widget_nummerierung2[i]):
                    widget_nummerierung[i] = int(widget_nummerierung[i])
                    widget_nummerierung2[i] = int(widget_nummerierung2[i])                    
                
            # Einloggen in starnet
            driver = webdriver.Chrome(ChromeDriverManager().install())
            url = 'https://starnet.labor28.sonichealthcare.de:8443/starnet-labor/login'
            driver.get(url)
            driver.maximize_window()
            # Bei Labor 28 anmelden
            login_name = dropdown_user.get()
            benutzername = driver.find_element('xpath','/html/body/div/div[4]/div[2]/div/div[1]/form/table/tbody/tr[1]/td[2]/input')
            benutzername.send_keys(login_name)
            passwort = driver.find_element('xpath','/html/body/div/div[4]/div[2]/div/div[1]/form/table/tbody/tr[2]/td[2]/input')
            passwort.send_keys(login_dict[login_name])
            login = driver.find_element('xpath','/html/body/div/div[4]/div[2]/div/div[1]/form/table/tbody/tr[3]/td[2]/input')
            login.click()
            driver.implicitly_wait(5)

            # Einträge generieren
            for i in range(number):
                # ersten Eintrag generieren
                auswahl_eintrag = driver.find_element('xpath','//*[@id="ID_ORDERS_OPEN_UPDATED_LIST_GRID"]/div[2]/div/table/tbody[2]/tr[1]/td[1]/div/div')
                auswahl_eintrag.click()
                wiederholen_button = driver.find_element('xpath', '//*[@id="ID_OPEN_UPDATED_ORDER_BUTTON_REDO"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
                wiederholen_button.click()
                patient = driver.find_element('xpath', '//*[@id="OrderNavigationViewImplPATIENT"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
                patient.click()
                entsperren = driver.find_element('xpath', '//*[@id="ID_ORDER_PATIENT_EDIT"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td/div')
                entsperren.click()
                # Probenname eingeben
                vorname_input = driver.find_element('xpath', '//*[@id="ID_ORDER_PATIENT_CONTENT_GIVENNAME-input"]')
                vorname_input.clear()
                vorname_input.send_keys(widget_probenname_vor[i])
                nachname_input = driver.find_element('xpath', '//*[@id="ID_ORDER_PATIENT_CONTENT_FAMILYNAME-input"]')
                nachname_input.clear()
                nachname_input.send_keys(widget_probenname_nach[i]+'_'+str(widget_nummerierung[i]))
                time.sleep(1)
                # Parameter
                # zuerst Klick auf Seite
                random_click = driver.find_element('xpath', '//*[@id="ID_ORDER_PATIENT_EDIT"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td/div')
                random_click.click()
                time.sleep(1)
                tests = driver.find_element('xpath', '//*[@id="OrderNavigationViewImplANALYSES"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]')
                tests.click()
                driver.implicitly_wait(5)
                entfernen = driver.find_element('xpath', '//*[@id="ID_ORDER_ANALYSES_REMOVE_ALL_BUTTON"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td/div')
                entfernen.click()
                time.sleep(1)
                # alle Parameter in widget_parameter[i] hinzufügen
                for j in range(len(widget_parameter[i])):
                    if j == 0:
                        hinzufügen = driver.find_element('xpath', '//*[@id="ID_ORDER_ANALYSES_ADD_BUTTON"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td/div')
                        hinzufügen.click()
                    testsuche = driver.find_element('xpath', '//*[@id="ID_ADD_ANALYSIS_DIALOG_SEARCH_ANALYSIS_COMBOBOX-input"]')
                    testsuche.send_keys(widget_parameter[i][j])
                    testsuche.send_keys(Keys.ENTER)
                    time.sleep(1)
                    find_parameter = driver.find_element('xpath', f"//*[text()='{widget_parameter[i][j].strip()}']")
                    find_parameter.click()
                tests_hinzufügen = driver.find_element('xpath', '//*[@id="ID_ADD_ANALYSIS_DIALOG_CONFIRM_BUTTON"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td/div')
                tests_hinzufügen.click()
                driver.implicitly_wait(5)
                time.sleep(2)
                # Abrechnung
                abrechnung = driver.find_element('xpath', '//*[@id="OrderNavigationViewImplBILLING"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
                abrechnung.click()
                driver.implicitly_wait(5)
                time.sleep(2)
                table = driver.find_element('xpath', '//*[@id="ID_ORDER_BILLINGS_GRID"]/div[2]/div/table')
                def get_element_xpath(element):
                    script = """
                    function getElementXPath(elt) {
                        var path = '';
                        for (; elt && elt.nodeType == 1; elt = elt.parentNode) {
                            var idx = getElementIdx(elt);
                            var xname = elt.tagName;
                            if (idx > 1) xname += '[' + idx + ']';
                            path = '/' + xname + path;
                        }
                        return path;
                    }

                    function getElementIdx(elt) {
                        var count = 1;
                        for (var sib = elt.previousSibling; sib; sib = sib.previousSibling) {
                            if (sib.nodeType == 1 && sib.tagName == elt.tagName) count++;
                        }
                        return count;
                    }

                    return getElementXPath(arguments[0]);
                    """
                    xpath = driver.execute_script(script, element)
                    return xpath

                # Find all input elements on the page
                input_elements = driver.find_elements('xpath', "//input")
                # Retrieve the XPath of each input element
                privat_input = []
                for element in input_elements:
                    xpath = get_element_xpath(element)
                    input_element = driver.find_element('xpath', xpath)
                    input_text = input_element.get_attribute("value")
                    if input_text == 'Privat':
                        privat_input.append(xpath)
                for el_xpath in privat_input:       
                    privat = driver.find_element('xpath', el_xpath)
                    privat.click()
                    privat.send_keys(Keys.DOWN)
                    privat.send_keys(Keys.ENTER)
                    time.sleep(1)
                # Fertig
                fertig = driver.find_element('xpath', '//*[@id="ID_ORDER_CONTENT_BUTTON_SAVE_AND_CLOSE"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
                fertig.click()
                driver.implicitly_wait(5)
                time.sleep(2)

                # Anzahl der zu generierenden Einträge aus widget_nummerierung und widget_nummerierung2
                if type(widget_nummerierung[i]) == int:
                    entries_nr = widget_nummerierung2[i] - widget_nummerierung[i]
                    # Einträge wiederholen, wobei nur der Probenname verändert wird
                    for l in range(entries_nr):
                        probe = driver.find_element('xpath', '//*[@id="ID_ORDERS_OPEN_UPDATED_LIST_GRID"]/div[2]/div/table/tbody[2]/tr[1]/td[1]/div/div')
                        probe.click()
                        # auf Wiederholen klicken
                        probe_bearbeiten = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div/div/div[4]/div/div/div[5]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
                        probe_bearbeiten.click()
                        driver.implicitly_wait(5)
                        time.sleep(1)
                        # Klick auf Patient
                        patient = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div[2]/div/div[3]/div/div[2]/div/div[8]/div/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
                        patient.click()
                        # Patientendaten entsperren
                        entsperren = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div[2]/div/div[1]/div/div/div[2]/div[1]/div/div[1]/div/div[1]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td/div')
                        entsperren.click()
                        driver.implicitly_wait(5)            
                        nachname_input = driver.find_element('xpath', '//*[@id="ID_ORDER_PATIENT_CONTENT_FAMILYNAME-input"]')
                        nachname_input.clear()
                        nachname_input.send_keys(widget_probenname_nach[i]+'_'+str(widget_nummerierung[i]+1+l))            
                        random_click = driver.find_element('xpath', '//*[@id="ID_ORDER_PATIENT_EDIT"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td/div')
                        random_click.click()
                        time.sleep(1)
                        random_click.click()
                        time.sleep(1)
                        fertig = driver.find_element('xpath', '//*[@id="ID_ORDER_CONTENT_BUTTON_SAVE_AND_CLOSE"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
                        fertig.click()
                        driver.implicitly_wait(5)
                        time.sleep(2)
                else:
                    entries_nr = []
                    startnr1 = int(widget_nummerierung[i].split('.')[0])
                    startnr2 = int(widget_nummerierung[i].split('.')[-1])
                    endnr1 = int(widget_nummerierung2[i].split('.')[0])
                    endnr2 = int(widget_nummerierung2[i].split('.')[-1])
                    diff1 = endnr1 - startnr1
                    diff2 = endnr2 - startnr2
                    
                    for i1 in range(diff1+1):
                        entries_nr.append(str(startnr1+i1))

                    entries_nr_f = []
                    for ie in range(len(entries_nr)):
                        entries = []
                        for i2 in range(diff2+1):
                            entries.append(entries_nr[ie]+'.'+str(i2+1))
                        entries_nr_f.append(entries)

                    entries_nr_f = [item for sublist in entries_nr_f for item in (sublist if isinstance(sublist, list) else [sublist])]
                    entries_nr_f = entries_nr_f[1:]
                    for nr in entries_nr_f:
                        probe = driver.find_element('xpath', '//*[@id="ID_ORDERS_OPEN_UPDATED_LIST_GRID"]/div[2]/div/table/tbody[2]/tr[1]/td[1]/div/div')
                        probe.click()
                        # auf Wiederholen klicken
                        probe_bearbeiten = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div/div/div[4]/div/div/div[5]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
                        probe_bearbeiten.click()
                        driver.implicitly_wait(5)
                        time.sleep(1)
                        # Klick auf Patient
                        patient = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div[2]/div/div[3]/div/div[2]/div/div[8]/div/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
                        patient.click()
                        # Patientendaten entsperren
                        entsperren = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div[2]/div/div[1]/div/div/div[2]/div[1]/div/div[1]/div/div[1]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td/div')
                        entsperren.click()
                        driver.implicitly_wait(5)            
                        nachname_input = driver.find_element('xpath', '//*[@id="ID_ORDER_PATIENT_CONTENT_FAMILYNAME-input"]')
                        nachname_input.clear()
                        nachname_input.send_keys(widget_probenname_nach[i]+'_'+nr)            
                        random_click = driver.find_element('xpath', '//*[@id="ID_ORDER_PATIENT_EDIT"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td/div')
                        random_click.click()
                        time.sleep(1)
                        random_click.click()
                        time.sleep(1)
                        fertig = driver.find_element('xpath', '//*[@id="ID_ORDER_CONTENT_BUTTON_SAVE_AND_CLOSE"]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
                        fertig.click()
                        driver.implicitly_wait(5)
                        time.sleep(2)
                    
                    
            fertig_ein = Label(root_run, text = 'Fertig! #TeamDodo',
                               bg = '#eeeee4', fg = '#869287', font=('Ink free',10,'bold'))
            fertig_ein.grid(row=30,column=1)
            
        # Leerzeile
        leer = Label(root_run, text = '', bg = '#eeeee4')
        leer.grid(row=20, column=1)
        go = Button(root_run, text = 'Go Dodo', bg = '#869287', font=('Ink free',12,'bold'),
                   command = go_dodo)
        go.grid(row=21, column=1)    
        
    dropdown.bind("<<ComboboxSelected>>", on_select)
    
    
    # Oder man entscheidet sich für die alte Methode, bei der man eine Excel-Datei hochlädt   
    
    # Funktion für Labor 28 Einträge, wenn man eine Excel-Tabelle hochladen möchte
    def run_excel():
        global file_path
        file_path= filedialog.askopenfilename(title = "Datei auswählen")
        #excel Datei einlesen
        df = pd.read_excel(file_path)
        liste_namen = list(df['Nachname'])
        driver = webdriver.Chrome(ChromeDriverManager().install())
        url = 'https://starnet.labor28.sonichealthcare.de:8443/starnet-labor/login'
        driver.get(url)
        driver.maximize_window()
        # Bei Labor 28 anmelden
        login_name = dropdown_user.get()
        benutzername = driver.find_element("xpath",'/html/body/div/div[4]/div[2]/div/div[1]/form/table/tbody/tr[1]/td[2]/input')
        benutzername.send_keys(login_name)
        passwort = driver.find_element("xpath",'/html/body/div/div[4]/div[2]/div/div[1]/form/table/tbody/tr[2]/td[2]/input')
        passwort.send_keys(login_dict[login_name])
        login = driver.find_element("xpath",'/html/body/div/div[4]/div[2]/div/div[1]/form/table/tbody/tr[3]/td[2]/input')
        login.click()
        driver.implicitly_wait(5)
        # Schleife, um Eintrag für alle Proben zu erstellen
        for i in range(len(liste_namen)):
            time.sleep(3)
            probe = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div/div/div[1]/div/div/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/table/tbody[2]/tr[1]/td[1]/div/div')
            probe.click()
            # auf Wiederholen klicken
            probe_bearbeiten = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div/div/div[4]/div/div/div[5]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
            probe_bearbeiten.click()
            driver.implicitly_wait(5)
            # Klick auf Patient
            patient = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div[2]/div/div[3]/div/div[2]/div/div[8]/div/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
            patient.click()
            # Patientendaten entsperren
            entsperren = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div[2]/div/div[1]/div/div/div[2]/div[1]/div/div[1]/div/div[1]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td/div')
            entsperren.click()
            driver.implicitly_wait(5)
            # Probenname eingeben
            nachname = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div[2]/div/div[1]/div/div/div[2]/div[1]/div/div[2]/div/div/div/div/div[3]/div[2]/div/div/div[5]/div[1]/div/div/input')
            nachname.clear()
            nachname.send_keys(liste_namen[i])
            speichern = driver.find_element("xpath",'/html/body/div[2]/div/div[2]/div[2]/div[1]/div/div[3]/div/div/div[2]/div/div[2]/div/div[4]/div/div/div[2]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[2]/div')
            speichern.click()
            time.sleep(3)

        fertig_excel = Label(root_run, text = 'Fertig! #TeamDodo',
                             bg = '#eeeee4', fg = '#869287', font=('Ink free',10,'bold'))
        fertig_excel.grid(row=30,column=1)
        
    excel_text = Label(root_run, text = '\n2. Methode: Wähle die Excel-Datei mit den Probennamen aus\n',
                       bg = '#eeeee4', font=('Ink free',12,'bold'))
    excel_text.grid(row=22, column=1, columnspan=5)
    browse = Button(root_run, text = 'Browse', bg = '#869287', font=('Ink free',12,'bold'),
                command = run_excel)
    browse.grid(row=23, column=1) 


# In[27]:


# Funktion für pdf Dateien (Rechnungen)
def rechnungen_func():
    global file_path
    file_path = filedialog.askopenfilename(title = "Datei auswählen")
    try:
        excel = pd.read_excel(file_path)
    except:
        excel = pd.read_csv(file_path, sep=';', encoding='latin-1')

    if (excel.index[0]==7122) | (excel['BETRAG'].isnull().all()):
        col_names = list(excel.columns)
        excel = excel.reset_index()
        excel = excel.iloc[:,:-1]
        excel.columns = col_names

    excel = excel[excel['BETRAG'].notna()]
    id_suche = list(excel['ExterneID'])
    id_suche = list(set(id_suche))
    id_suche = [int(el) for el in id_suche if str(el) != 'nan']
    probenname = np.empty((len(id_suche),0)).tolist()
    
    login_name = dropdown_user.get()
    
    driver = webdriver.Chrome(ChromeDriverManager().install())
    url = 'https://starnet.labor28.sonichealthcare.de:8443/starnet-labor/login'
    driver.get(url)
    driver.maximize_window()
    # Bei Labor 28 anmelden
    benutzername = driver.find_element("xpath",'/html/body/div/div[4]/div[2]/div/div[1]/form/table/tbody/tr[1]/td[2]/input')
    benutzername.send_keys(login_benutzername[login_name])
    passwort = driver.find_element("xpath",'/html/body/div/div[4]/div[2]/div/div[1]/form/table/tbody/tr[2]/td[2]/input')
    passwort.send_keys(login_passwort[login_name])
    login = driver.find_element("xpath",'/html/body/div/div[4]/div[2]/div/div[1]/form/table/tbody/tr[3]/td[2]/input')
    login.click()
    driver.implicitly_wait(5)

    wait = WebDriverWait(driver, 10)
    wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
    befundet = driver.find_element("xpath",'//*[@id="REPORT_ALL"]')
    befundet.click()
    wait = WebDriverWait(driver, 10)
    wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

    for i in range(len(id_suche)):
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
        time.sleep(1)
        suche = driver.find_element("xpath", '//*[@id="ID_COMMONS_TOOLBAR_SEARCH_FIELD-input"]')
        suche.clear()
        suche.send_keys(id_suche[i])
        suche.send_keys(Keys.ENTER)
        time.sleep(1)
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
        aktualisieren = driver.find_element("xpath", '//*[@id="TAB_BACKGROUND"]/div[2]/div/div/div/div[1]/div/div/div/div[2]/div[2]/div[1]/div/div[2]/div/div/div/div[10]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody')
        aktualisieren.click()
        wait = WebDriverWait(driver, 10)
        time.sleep(2)
        xpath = '//*[@id="ID_ORDERS_REPORT_ALL_LIST_GRID"]/div[2]/div/table/tbody[2]/tr/td[11]/div/div/b'
        wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        
        for attempt in range(3):        
            try:
                text_suche = driver.find_element("xpath", xpath)
                text = text_suche.text
                probenname[i] = text.split('+')[0]
                break
            except Exception as e:
                time.sleep(1)
        
    driver.close()  
    
    # Zugreifen auf Produktionsübersichtstabelle, um Info zu IP-Nummern zu extrahieren
    from datetime import date
    current_year = date.today().year
    datei = pd.ExcelFile('O:\Produktion\Übersicht\Produktnummern.xlsx')
    ip_thisyear = pd.read_excel(datei, '{}'.format(current_year))
    ip_thisyear = ip_thisyear[['Lot \n(add .1.1)', 'Operator', 'Description', 'Customer']]

    # manche IP-Nummern sind vom vorherigen Jahr -> anderer Reiter in der Excel-Tabelle
    ip_oneyearago = pd.read_excel(datei, '{}'.format(current_year-1))
    ip_oneyearago = ip_oneyearago[['Lot \n(add .1.1)', 'Operator', 'Description', 'Customer']]
    ip_oneyearago.columns = ['Lot \n(add .1.1)', 'Operator', 'Description', 'Customer']
    ip_nummern = pd.concat([ip_oneyearago, ip_thisyear]).dropna().reset_index().drop('index', axis = 1)   
    
    # Spalte Proben ID
    probenname_f = probenname.copy()
    muster_id = r'100\d{9}|100\d{6}|\d{4}-\d{2}|\d{4}-\d{3}'
    for i in range(len(probenname_f)):
        match = re.finditer(muster_id, probenname_f[i])
        if match:
            probenname_f[i] = probenname_f[i].split(',')[0].split('.')[0].split('_')[0].split(' ')[0]
        if 'IP' in probenname_f[i]:
            probenname_f[i] = probenname_f[i].replace('IP', '')
        probenname_f[i] = probenname_f[i].strip()
    
    # dictionary auftragsnummer: probenname
    dict_probenname = dict(zip(id_suche, probenname_f))
    dict_probenname_f = dict(zip(probenname, probenname_f))

    # neues df
    df = excel.copy()
    df = df[['ExterneID', 'RECHZEILE', 'BETRAG']]
    df['ExterneID'] = df['ExterneID'].map(dict_probenname)
    df.columns = ['Proben ID', 'Parameter', 'Preis/Messung [€]']
    
    # für jede Probe Abteilung zuordnen
    if login_name == 'F&E':
        abteilung = ['Forschung & Entwicklung']*len(df)
    elif login_name == 'P':
        abteilung = ['Produktion']*len(df)
    elif login_name == 'QS':
        abteilung = ['Qualitätssicherung']*len(df)
    else:
        abteilung = ['Datenmanagement/Probenmanagement']*len(df)
    df['Abteilung'] = abteilung
                
    # unique ip Liste zum Suchen in der Produktionsübersichttabelle
    unique = list(set(probenname_f))
    
    index_df_ip_extra = []
    for un in unique:
        for i in range(len(ip_nummern)):
            if un in ip_nummern.iloc[i,0]:
                index_df_ip_extra.append(i)
                break
    
    df_ip_extra = ip_nummern.iloc[index_df_ip_extra,:]
    df_ip_extra.columns = ['Sample ID', 'Operator', 'Description', 'Customer']

    # dictionary ip: description
    keys = [x.split('IP')[-1].strip() for x in list(df_ip_extra['Sample ID'])]
    values = list(df_ip_extra['Description'])
    dict_descr = dict(zip(keys, values))
    # same for customer column
    values_c = list(df_ip_extra['Customer'])
    dict_cust = dict(zip(keys, values_c))

    # Spalte mit Operator der Vermessung hinzufügen
    fe_kuerzel = ['LQu', 'MBu', 'HZe', 'ASe', 'LHü', 'LHue']
    p_kuerzel = ['DWe', 'CSi', 'AGr', 'DVö']
    qs_kuerzel = ['STi', 'LKä', 'JMi']
    
    operator = ['']*len(df)
    description = ['-']*len(df)
    kuerzel = fe_kuerzel + p_kuerzel + qs_kuerzel
    for i in range(len(df)):
        for k in kuerzel:
            if k in df.iloc[i, 0]:
                operator[i] = k
                break
    df['Operator (Messung)'] = operator
    
    df['Proben ID'] = df['Proben ID'].map(dict_probenname_f)
    
    # add description to df_ip
    for i in range(len(df)):
        for ip in dict_descr:
            if ip in df.iloc[i, 0]:
                description[i] = dict_descr[ip]
    df['Description'] = description
    
    # restl. Infos ergänzen    
    df['Customer'] = df['Proben ID'].map(dict_cust)
    
    # fill na, da manche IP-Nummern nicht in der Übersichtstabelle sind
    df = df.replace(np.nan, '-')
    
    # Dataframe ergänzen
    index_df = [list(df['Proben ID']),
                list(df['Abteilung']),
                list(df['Description']),
                list(df['Customer']),
                list(df['Operator (Messung)']),
                list(df['Parameter']),
                list(df['Preis/Messung [€]'])
               ]

    index = []
    for i in index_df:
        index.append(i)

    tuples = list(zip(*index))
    multi_index = pd.MultiIndex.from_tuples(tuples, names=['Proben ID', 'Abteilung', 'Description', 'Customer', 'Operator (Messung)', 'Parameter','Preis/Messung [€]'])
    liste = []
    df2 = pd.DataFrame(liste*len(df),index=multi_index)
    anzahl = [1]*len(df)
    df2['Anzahl Messungen/Parameter'] = anzahl
    df3 = df2.groupby(['Abteilung', 'Proben ID', 'Description', 'Customer', 'Operator (Messung)','Parameter','Preis/Messung [€]'])['Anzahl Messungen/Parameter'].count()
    df3 = pd.DataFrame(df3)

    df_final = df3.reset_index()
    df_new = df_final.set_index(['Abteilung','Proben ID', 'Description', 'Customer', 'Operator (Messung)'])
    # Wenn kein Preis angegeben wurde, wird 0 eingetragen
    df_new['Preis/Messung [€]'] = df_new['Preis/Messung [€]'].replace("-", 0)
    for i in range(len(df_new)):
        if ',' in str(df_new['Preis/Messung [€]'][i]):
            df_new['Preis/Messung [€]'][i] = float(str(df_new['Preis/Messung [€]'][i]).replace(',', '.'))
    df_new['Preis/Messung [€]'] = df_new['Preis/Messung [€]'].map(float)
        
#     if (df_new['Preis/Messung [€]'].astype(float) != df_new['Preis/Messung [€]']).all():
#         df_new['Preis/Messung [€]'] = df_new['Preis/Messung [€]'].str.replace(',', '.').map(float)

    platzhalter = [1]*len(df_new)
    df_new['Anzahl Messungen*Preis [€]'] = platzhalter
    df_new['Anzahl Messungen*Preis [€]'] = df_new['Anzahl Messungen/Parameter']*df_new['Preis/Messung [€]']
    
    writer = pd.ExcelWriter('Labor 28 Rechnungen Ergebnis.xlsx', engine='xlsxwriter')
    df_new.to_excel(writer, sheet_name = 'Übersicht')
    writer.close()
    
    # insert to excel
    excel = auswertung_pfad
    workbook = openpyxl.load_workbook(excel)
    worksheet = workbook['Labor 28']
    
    def insert_to_auswertung(df, start_row, start_col):
        """ füllt die Auswertung-Datei für entsprechende Abteilung aus
        """
        
        num_rows, num_cols = df.shape
        data = df.values.tolist()

        for r_idx, row in enumerate(data):
            for c_idx, cell_value in enumerate(row):
                worksheet.cell(row=start_row + r_idx, column=start_col + c_idx, value=cell_value)
    
    if login_name == 'DM':
        df_dm = df_new.loc['Datenmanagement/Probenmanagement',:].reset_index()
        df_dm = df_dm.groupby(['Parameter','Preis/Messung [€]'])['Anzahl Messungen/Parameter'].sum()
        df_dm = df_dm.reset_index()
        para_df = list(df_dm['Parameter'])

        dict_para_dm = {'ANCA-/p-ANCA-Titer':'c-ANCA (Titer)',
                        'CCP-Ak':'a-CCP',
                        'Calprotectin':'Elastase oder Calprotectin',
                        'Elastase (pankr.) i. St.': 'Elastase oder Calprotectin',
                        'DNS AK Crithidien-IFT-Titer': 'DNA AK Crithidien IFT Titer', 
                        'DNS AK Crithidien-IFT': 'DNA AK Crithidien IFT',
                        'Immunglobulin G (IgG)': 'IgG',
                        'SS-A/Ro52-Autoantikörper':'SS-A (Ro60 od. Ro52)',
                        'SS-A/Ro60-Autoantikörper':'SS-A (Ro60 od. Ro52)',
                        'Transglutaminase IgA-Ak':'Transglutaminase IgA',
                        'ADAMTS-13-Antigen':'ADAMTS-13-Ag',
                        'ADAMTS-13-Antikörper':'ADAMTS-13-AK',
                        'ADAMTS-13-Aktivität':'ADAMTS-13-Aktivität',
                        'TRAK':'a-TSH-Rezeptor (TRAK)',
                        'Immunglobulin E, gesamt (IgE)': 'IgE, gesamt',
                        'Myeloperoxidase-AK': 'MPO-AK',
                        'Centromer-B-Autoantikörper':'Centromer-A-AAk/ Centromer-B-AAk',
                        'Centromer-B-Autoantikörper':'Centromer-A-AAk/ Centromer-B-AAk',
                        'Acetylcholinrezeptor Autoantikörper':'Acetylcholinrezeptor AK/ ACREP',
                        'mikrosomale TPO AK':'a-TPO (Mikrosomale Peroxidase AK)',
                        'Thyreoglobulin AK':'a-Tg (Thyreoglobulin AK)',
                        'Borrelien AK (IgG + IgM)':'Borrelien IgG/IgM AK',
                        'Borrelien Blot (IgG)':'Borrelia Burg. Blot IgG',
                        'DNS Autoantikörper (ds-DNS)':'DNA-AK',
                        'ENA-Ak Screening':'ENA-AK Screening',
                        'GAD-AK (Glutamat-Decarboxylase)':'GAD-AK',
                        'AAK g. glomeruläre Basalmembran':'glomeruläre Basalmembran AK (GBM)',
                        'HbA 1c':'HbA1c',
                        'Helicobacter pylori Antigen (Stuhl)':'H. pylori Antigen',
                        'Inselzell Autoantikörper (ICA)':'Inselzell AAK (ICA)',
                        'Tyrosin-Phosphastase (IA2)- AK(IA2)':'Inselzell-Ag-2-AK',
                        'Insulin Autoantikörper (IAA)':'Insulin AAK (IAA)',
                        'HBV-DNA':'PCR (HBV)',
                        'HCV-RNA':'PCR (HCV)',
                        'HIV-1 PCR (Viral-load, ultrasensitiv)':'PCR (HIV/HCV/HBV)',
                        'PR3-ANCA':'Proteinase 3-AK (PR3)',
                        'IgA-Rheumafaktor':'RF IgA',
                        'IgG-Rheumafaktor':'RF IgG',
                        'IgM-Rheumafaktor':'RF IgM',
                        'SmD-Autoantikörper':'SmD',
                        'Transglutaminase IgA-AK':'Transglutaminase IgA',
                        'TSH (sensitiv)':'TSH'
                        }

        for i in range(len(para_df)):
            allergen = ' (Allergen)'
            if allergen in para_df[i]:
                para_df[i] = para_df[i].split(allergen)[0]
            elif 'RF' in para_df[i]:
                para_df[i] = 'RF'
            elif 'Toxo' in para_df[i]:
                para_df[i] = 'Toxo'

        for i in range(len(para_df)):
            if para_df[i] in dict_para_dm.keys():
                para_df[i] = dict_para_dm[para_df[i]]    

        df_dm['Parameter'] = para_df
        df_dm = df_dm.groupby(['Parameter', 'Preis/Messung [€]'])['Anzahl Messungen/Parameter'].sum()
        df_dm = pd.DataFrame(df_dm).reset_index()

        #auswertung = pd.read_excel('Auswertung von Laborrechnungen.xlsx')    

        auswertung_dm = auswertung.iloc[5:,13:16]
        auswertung_dm.columns = ['Bezeichnung', 'Preis [€]', 'Anzahl']
        auswertung_dm = auswertung_dm.reset_index().drop('index', axis=1)

        # Länge von Tabelle in Excel
        len_auswertung_dm = len(auswertung_dm) 

        for i in range(len(df_dm)):
            for ii in range(len(auswertung_dm)):
                if df_dm.iloc[i,0] == auswertung_dm.iloc[ii,0]:
                    auswertung_dm.iloc[ii,2]=df_dm.iloc[i,2]
                    auswertung_dm.iloc[ii,1]=df_dm.iloc[i,1]
        auswertung_dm = auswertung_dm.dropna().reset_index().drop('index', axis=1)
        auswertung_dm.columns = ['Parameter', 'Preis/Messung [€]', 'Anzahl Messungen/Parameter']

        para_dm = list(auswertung_dm['Parameter'])
        para_df_dm = list(df_dm['Parameter'])
        add_para = [p for p in para_df_dm if p not in para_dm]

        for p in add_para:
            df_dm_sub = df_dm[df_dm['Parameter']==p]
            auswertung_dm = pd.concat([auswertung_dm, df_dm_sub], ignore_index=True)

        # leere Zeilen in auswertung_dm einfügen, damit diese in Excel kopiert werden
        empty_rows = len_auswertung_dm-len(auswertung_dm)
        empty_df = pd.DataFrame()
        empty_df['Parameter'] = [' ']
        empty_df['Preis/Messung [€]'] = ['']
        empty_df['Anzahl Messungen/Parameter'] = ['']
        for i in range(empty_rows):
            auswertung_dm = pd.concat([auswertung_dm, empty_df], ignore_index=True)
        
        insert_to_auswertung(auswertung_dm, 7, 14)
        
    
    # Auswertung-Tabelle ausfüllen für F&E
    elif login_name == 'F&E':
        df_fe = df_new.loc['Forschung & Entwicklung',:].reset_index()

        for i in range(len(df_fe)):
            if df_fe.iloc[i,1] == '-':
                df_fe.iloc[i,1] = df_fe.iloc[i,4]
            if df_fe.iloc[i,3] != '':
                df_fe.iloc[i,1] = df_fe.iloc[i,1] + ', ' + df_fe.iloc[i,3]
            
        df_fe = df_fe.groupby(['Description','Preis/Messung [€]'])['Anzahl Messungen/Parameter'].sum()
        df_fe = df_fe.reset_index()
        df_fe['Summe'] = df_fe['Preis/Messung [€]']*df_fe['Anzahl Messungen/Parameter']
        
        # Unterscheiden, wenn mehrere Parameter oder nur einer für Projekt vermessen wurde:
        count = df_fe['Description'].value_counts()
        count = pd.DataFrame(count).reset_index()
        one_meas = []
        for i in range(len(count)):
            if count.iloc[i,1] == 1:
                one_meas.append(count.iloc[i,0])

        df_fe_one = df_fe[df_fe['Description'].isin(one_meas)].reset_index().drop('index', axis=1).drop('Summe', axis=1)
        df_fe_one.columns = ['Bezeichnung', 'Preis', 'Anzahl']

        df_fe_more = df_fe[~df_fe['Description'].isin(one_meas)].reset_index().drop('index', axis=1)
        df_fe_more = pd.DataFrame(df_fe_more.groupby('Description').sum()['Summe']).reset_index()
        df_fe_more['Anzahl'] = [1]*len(df_fe_more)
        df_fe_more.columns = ['Bezeichnung', 'Preis', 'Anzahl']
        
        df_fe = pd.concat([df_fe_one, df_fe_more])
        
        insert_to_auswertung(df_fe, 7, 20)
    
    elif login_name == 'P':
        # Produktion
        df_p = df_new.loc['Produktion',:].reset_index()
        
        dict_new_desc = dict(zip(list(df_p['Proben ID']), list(df_p['Description'])))
        dict_new_desc2 = list(dict_new_desc.values())
        dict_new_desc2 = list(set([el for el in dict_new_desc2 if el != '-']))
        dict_new_desc3 = np.empty((len(dict_new_desc2),0)).tolist()

        liste_values = list(dict_new_desc.values())
        liste_keys = list(dict_new_desc.keys())
        for i in range(len(dict_new_desc2)):
            for j in range(len(liste_values)):
                if dict_new_desc2[i] == liste_values[j]:
                    dict_new_desc3[i].append(liste_keys[j])

        for i in range(len(dict_new_desc3)):
            dict_new_desc2[i] = dict_new_desc2[i] + ' ' + str(dict_new_desc3[i])
            dict_new_desc2[i] = dict_new_desc2[i].replace("'", "").replace('[', '(').replace(']', ')')

        for i in range(len(df_p)):
            for d in dict_new_desc2:
                if df_p.iloc[i,0] in d:
                    df_p.iloc[i,1] = d
            if df_p.iloc[i,1] == '-':
                df_p.iloc[i,1] = df_p.iloc[i,4]        
                
        # Operator ergänzen
        for i in range(len(df_p)):
            if df_p.iloc[i,1] == '-':
                df_p.iloc[i,1] = df_p.iloc[i,4]
            if df_p.iloc[i,3] != '':
                df_p.iloc[i,1] = df_p.iloc[i,1] + ', ' + df_p.iloc[i,3]            
            

        liste_bezeichnung = list(df_p['Description'])

        for i in range(len(liste_bezeichnung)):
            muster = r'RV\d{3} L\d{1}'
            match = re.search(muster, liste_bezeichnung[i])
            if match:
                liste_bezeichnung[i] = match.group()
            muster2 = r'RV \d{3} L\d{1}'
            match = re.search(muster2, liste_bezeichnung[i])
            if match:
                liste_bezeichnung[i] = match.group()

        df_p['Description'] = liste_bezeichnung

        df_p = df_p.groupby(['Description','Preis/Messung [€]'])['Anzahl Messungen/Parameter'].sum()
        df_p = df_p.reset_index()
        df_p['Summe'] = df_p['Preis/Messung [€]']*df_p['Anzahl Messungen/Parameter']

        # Unterscheiden, wenn mehrere Parameter oder nur einer für Projekt vermessen wurde:
        count = df_p['Description'].value_counts()
        count = pd.DataFrame(count).reset_index()
        one_meas = []
        for i in range(len(count)):
            if count.iloc[i,1] == 1:
                one_meas.append(count.iloc[i,0])

        df_p_one = df_p[df_p['Description'].isin(one_meas)].reset_index().drop('index', axis=1).drop('Summe', axis=1)
        df_p_one.columns = ['Bezeichnung', 'Preis', 'Anzahl']

        df_p_more = df_p[~df_p['Description'].isin(one_meas)].reset_index().drop('index', axis=1)
        df_p_more = pd.DataFrame(df_p_more.groupby('Description').sum()['Summe']).reset_index()
        df_p_more['Anzahl'] = [1]*len(df_p_more)
        df_p_more.columns = ['Bezeichnung', 'Preis', 'Anzahl']
        
        df_p = pd.concat([df_p_one, df_p_more])
        
        insert_to_auswertung(df_p, 7, 8)
    
    # für QS
    elif login_name == 'QS':
        df_qs = df_new.loc['Qualitätssicherung',:].reset_index()
                
        dict_new_desc = dict(zip(list(df_qs['Proben ID']), list(df_qs['Description'])))
        dict_new_desc2 = list(dict_new_desc.values())
        dict_new_desc2 = list(set([el for el in dict_new_desc2 if el != '-']))
        dict_new_desc3 = np.empty((len(dict_new_desc2),0)).tolist()

        liste_values = list(dict_new_desc.values())
        liste_keys = list(dict_new_desc.keys())
        for i in range(len(dict_new_desc2)):
            for j in range(len(liste_values)):
                if dict_new_desc2[i] == liste_values[j]:
                    dict_new_desc3[i].append(liste_keys[j])

        for i in range(len(dict_new_desc3)):
            dict_new_desc2[i] = dict_new_desc2[i] + ' ' + str(dict_new_desc3[i])
            dict_new_desc2[i] = dict_new_desc2[i].replace("'", "").replace('[', '(').replace(']', ')')

        for i in range(len(df_qs)):
            for d in dict_new_desc2:
                if df_qs.iloc[i,0] in d:
                    df_qs.iloc[i,1] = d
            if df_qs.iloc[i,1] == '-':
                df_qs.iloc[i,1] = df_qs.iloc[i,0]
                    
        # Operator ergänzen
        for i in range(len(df_qs)):
            if df_qs.iloc[i,3] != '':
                df_qs.iloc[i,1] = df_qs.iloc[i,1] + ', ' + df_qs.iloc[i,3]

        df_qs = pd.DataFrame(df_qs.groupby('Description').sum()['Anzahl Messungen*Preis [€]']).reset_index()
        df_qs['Anzahl'] = [1]*len(df_qs)
        
        insert_to_auswertung(df_qs, 7, 2)
                       
    
    workbook.save(excel)
    
    
    tada = Label(root2, text = '\nFertig! #TeamDodo kriegt von der QS einen Kuchen',
                 font=('Ink free',10,'bold'), bg = '#eeeee4', fg = '#869287').pack()            
    zurück = Button(root2, text = 'Zurück', bg = '#869287', font=('Ink free',12,'bold'), command = goback).pack()

    
    
    
####### SYNLAB ########
def synlab():
    
    global files, df
    files = filedialog.askopenfilenames(title = "Datei auswählen")
    files = list(files)
    
    df = pd.DataFrame()
    df['Proben ID'] = []

    # read_function definieren
    def read_function(file):
        global df
        excel = pd.read_excel(file, skiprows=1)[:-2]
        excel = excel.dropna(subset=['LOT#'])
        parameter = list(excel.columns[2:-1])
        # Probennamen
        muster_id = r'100\d{9}|100\d{6}|\d{4}-\d{2}|\d{4}-\d{3}'
        proben = list(excel['LOT#'])
        for i in range(len(proben)):
            match = re.search(muster_id, proben[i])
            if match:
                proben[i] = match.group()
        excel['LOT#'] = proben
        # Excel bearbeiten
        excel = excel.drop('Matrix', axis=1)
        column_names = ['Probe'] + parameter + ['Operator']
        excel.columns = column_names
        # Parameter identifizieren
        for i in range(len(excel)):
            for j in range(len(excel.columns)):
                if (excel.iloc[i,j] == 'X') | (excel.iloc[i,j] == 'x'):
                    excel.iloc[i,j] = excel.columns[j]
        parameter = np.empty((len(excel),0)).tolist()
        for i in range(len(excel)):
            for j in range(1,len(excel.columns)-1):
                parameter[i].append(excel.iloc[i,j])
        for i in range(len(parameter)):
            parameter[i] = [p for p in parameter[i] if p != '-' and str(p) != 'nan']
        # Operator
        kuerzel = list(excel['Operator'])

        proben_final = np.empty((len(parameter),0)).tolist()
        operator = np.empty((len(parameter),0)).tolist()
        for i in range(len(proben_final)):
            proben_final[i] = np.empty((len(parameter[i]),0)).tolist()
            operator[i] = np.empty((len(parameter[i]),0)).tolist()

        for i in range(len(proben_final)):
            for ii in range(len(proben_final[i])):
                proben_final[i][ii] = proben[i]
                operator[i][ii] = kuerzel[i]

        for i in range(len(parameter)):
            index_drop = []
            for j in range(len(parameter[i])):
                if 'vermessen' in parameter[i][j]:
                    index_drop.append(j)
        if index_drop != []:
            proben_final[i] = [proben_final[i][l] for l in range(len(proben_final[i])) if l not in index_drop]
            parameter[i] = [parameter[i][l] for l in range(len(parameter[i])) if l not in index_drop]
            operator[i] = [operator[i][l] for l in range(len(operator[i])) if l not in index_drop]
            
        for j in range(len(parameter[i])):
            for x in ['; \n', ';\n']:
                if x in parameter[i][j]:
                    insert_i = j+1
                    parameter[i].insert(insert_i, parameter[i][j].split(x)[-1])
                    parameter[i][j] = parameter[i][j].split(x)[0]
                    proben_final[i].insert(insert_i, proben_final[i][j])
                    operator[i].insert(insert_i, operator[i][j])
                    break
                    
        proben = [p for sublist in proben_final for p in sublist]
        parameter = [p for sublist in parameter for p in sublist]
        operator = [p for sublist in operator for p in sublist]

        # Abteilung zuordnen
        abt = ['-']*len(operator)
        fe_kuerzel = ['LQu', 'HZe', 'ASe', 'LHü', 'LHue', 'MBu']
        p_kuerzel = ['DVö', 'DVo', 'DVoe', 'CSi', 'DWe', 'AGr']
        qs_kuerzel = ['STi', 'LKä', 'LKa', 'LKae', 'SRe', '1SRe', 'JMi']

        for i in range(len(operator)):
            if operator[i] in fe_kuerzel:
                abt[i] = 'Forschung & Entwicklung'
            elif operator[i] in p_kuerzel:
                abt[i] = 'Produktion'
            elif operator[i] in qs_kuerzel:
                abt[i] = 'Qualitätssicherung'
            else:
                abt[i] = 'Datenmanagement/Probenmanagement'


        system = list(map(lambda x: x.split('(')[-1].split(')')[0], parameter))
        parameter = list(map(lambda x: x.split(' (')[0], parameter))
        for i in range(len(parameter)):
            if '/ ' in parameter[i]:
                parameter[i] = parameter[i].split('/ ')[-1]
            elif '/' in parameter[i]:
                parameter[i] = parameter[i].split('/')[-1]
            parameter[i] = parameter[i].partition('  ')[0]

        for i in range(len(system)):
            if 'im ' in system[i]:
                system[i] = system[i].split('im ')[-1].split(' ')[-1]

        excel = pd.DataFrame({'Proben ID': proben, 'Operator': operator, 'Parameter': parameter, 'Abteilung': abt, 'Gerät':system})

        df = pd.concat([df,excel], ignore_index=True)
    
    for i in range(len(files)):
        read_function(files[i])

    df['Anzahl'] = [1]*len(df)

    # Manchmal wird eine Mehrfachbestimmung beauftragt
    drop_row = []
    for i in range(len(df)-1):
        if 'Bestimmung' in df.iloc[i+1,2]:
            drop_row.append(i+1)
            muster = r'\d{1}'
            match = re.search(muster, df.iloc[i+1,2])
            if match:
                factor = match.group()
                df.iloc[i,-1] = df.iloc[i,-1]*factor
    if drop_row != []:
        df = df.drop(drop_row).reset_index().drop('index', axis=1)

    # Zugreifen auf Produktionsübersichtstabelle, um Info zu IP-Nummern zu extrahieren
    from datetime import date
    current_year = date.today().year
    datei = pd.ExcelFile('O:\Produktion\Übersicht\Produktnummern.xlsx')
    ip_thisyear = pd.read_excel(datei, '{}'.format(current_year))
    ip_thisyear = ip_thisyear[['Lot \n(add .1.1)', 'Operator', 'Description', 'Customer']]

    # manche IP-Nummern sind vom vorherigen Jahr -> anderer Reiter in der Excel-Tabelle
    ip_oneyearago = pd.read_excel(datei, '{}'.format(current_year-1))
    ip_oneyearago = ip_oneyearago[['Lot \n(add .1.1)', 'Operator', 'Description', 'Customer']]
    ip_oneyearago.columns = ['Lot \n(add .1.1)', 'Operator', 'Description', 'Customer']
    ip_nummern = pd.concat([ip_oneyearago, ip_thisyear]).dropna().reset_index().drop('index', axis = 1)   

    # unique ip Liste zum Suchen in der Produktionsübersichttabelle
    ip_liste = df['Proben ID']
    unique = list(set(ip_liste))

    ip_nummern['Lot \n(add .1.1)'] = ip_nummern['Lot \n(add .1.1)'].map(lambda x: x.replace('IP','').strip())
    df_ip_extra =ip_nummern.loc[ip_nummern['Lot \n(add .1.1)'].isin(unique)]
    df_ip_extra.columns = ['Sample ID', 'Operator', 'Description', 'Customer']

    # dictionary ip: description
    keys = list(df_ip_extra['Sample ID'])
    values = list(df_ip_extra['Description'])
    dict_descr = dict(zip(keys, values))

    # add description to df_ip
    df['Description'] = df['Proben ID'].map(dict_descr)

    # same for customer column
    values_c = list(df_ip_extra['Customer'])
    dict_cust = dict(zip(keys, values_c))

    # add description to df_ip
    df['Customer'] = df['Proben ID'].map(dict_cust)

    # fill na, da manche IP-Nummern nicht in der Übersichtstabelle sind
    df = df.replace(np.nan, '-')

    preise_path = 'O:\\Datenmanagement\\Befunde und Messsysteme\\Messsysteme\\1_Auftragslabore'
    preise_files = os.listdir(preise_path)
    suche_excel = 'Messsysteme und Preise aller Auftragslabore'

    try:
        for i in range(len(preise_files)):
            if suche_excel in preise_files[i]:
                file = preise_files[i]
                if file[:2] == '~$':
                    file = file[2:]
        preise = pd.read_excel(preise_path+'\\'+file, sheet_name='Synlab (ehem. W&T)', skiprows=2)
        preise = preise.iloc[::-1]
    except:
        preise_files = glob.glob('*.xlsx*')
        for i in range(len(preise_files)):
            if suche_excel in preise_files[i]:
                file = preise_files[i]
                if file[:2] == '~$':
                    file = file[2:]
        preise = pd.read_excel(preise_path+'\\'+file, sheet_name='Synlab (ehem. W&T)', skiprows=2)
        preise = preise.iloc[::-1]    

    df = df.groupby(['Abteilung', 'Proben ID', 'Operator', 'Description', 'Customer', 'Parameter', 'Gerät'])['Anzahl'].sum()
    df = pd.DataFrame(df.reset_index().set_index(['Abteilung', 'Operator', 'Description', 'Customer']))[['Proben ID', 'Parameter', 'Gerät', 'Anzahl']]
    df['Anzahl'] = df['Anzahl'].map(int)
        
    drop_rows = []
    for i in range(len(df)-1):
        if df.iloc[i,0] == df.iloc[i+1,0]:
            first_para = df['Parameter'][i]
            second_para = df['Parameter'][i+1]
            if (('HBs-Ag' in first_para) & ('HIV-1+2-Ak\nHCV-Ak' in second_para)) | (('HBs-Ag' in first_para) & ('HIV-1+2-Ak\nHCV-Ak' in second_para)):
                drop_rows.append(i+1)
                df.iloc[i,1] = 'Infektionsserologie Profil: HIV-AK, HBsAG, Anti-HCV'

    df = df.reset_index()
    df = df.drop(drop_rows)
    #df = df.set_index(['Abteilung', 'Proben ID', 'Description', 'Customer'])
    df = df.groupby(['Abteilung', 'Proben ID', 'Operator', 'Description', 'Customer', 'Parameter', 'Gerät', 'Anzahl']).count()
    df = pd.DataFrame(df.reset_index().set_index(['Abteilung', 'Proben ID', 'Operator', 'Description', 'Customer'])[['Parameter', 'Gerät', 'Anzahl']])

    # manchmal wird 'Gesamt T3' und 'Gesamt T4' zusammen vermessen
    # Schreibweise ist manchmal unterschiedlich
    for i in range(len(df)):
        if ('Gesamt T3' in df.iloc[i,0]) & ('Gesamt T4' in df.iloc[i,0]):
            df.iloc[i,0] = 'Gesamt T3, Gesamt T4'

    # Preise für jeden Parameter in einem Dictionary festhalten
    para_syst_dict = {}
    for i in range(len(df)):
        para_syst_dict[df['Parameter'][i]] = df['Gerät'][i]

    parameter = list(set(list(df['Parameter'])))
    # "anti" mit "a" ersetzen und für die Suche in die Parameter-Liste aufnehmen
    dict_anti = {}
    for i in range(len(parameter)):
        if 'anti' in parameter[i]:
            if ('TPO' not in parameter[i]) & ('Tg' not in parameter[i]):
                for j in list(preise['Parameter']):
                    if parameter[i] not in list(preise['Parameter']):
                        dict_anti[parameter[i]] = parameter[i].replace('anti', 'a')
                        parameter[i] = parameter[i].replace('anti', 'a')
                        break
    parameter_new = list(df['Parameter'])
    for i in range(len(parameter_new)):
        if parameter_new[i] in list(dict_anti.keys()):
            parameter_new[i] = dict_anti[parameter_new[i]]
    df['Parameter'] = parameter_new

    preise_dict = {'Infektionsserologie Profil: HIV-AK, HBsAG, Anti-HCV': 17.20, 'CMV-PCR': 4.85,
                  'Gesamt T3, Gesamt T4':14.58}


    def add_price(parameter):
        ''' argument: Parameter-Liste
        '''
        for i in range(len(parameter)):
            for ii in range(len(preise)):
                if parameter[i] in preise.iloc[ii,0]:
                    if (para_syst_dict[parameter[i]] in str(preise.iloc[ii,12])) | (para_syst_dict[parameter[i]] in str(preise.iloc[ii,11])):
                        preise_dict[parameter[i]] = preise.iloc[ii,4]
                        if np.isnan(preise_dict[parameter[i]]):
                            preise_dict[parameter[i]] = preise.iloc[ii-1,4]
                            break
                    if parameter[i] not in list(preise_dict.keys()):
                        preise_dict[parameter[i]] = preise.iloc[ii,4]
                        if np.isnan(preise_dict[parameter[i]]):
                            preise_dict[parameter[i]] = preise.iloc[ii-1,4]
                            break
        return(preise_dict)

    add_price(parameter)

    # manchmal werden IgG, IgM und IgA zusammen als ein Parameter angegeben
    # dabei handelt es sich um mehrere Messungen
    antibodies = ['IgG', 'IgA', 'IgM']
    extra_suche = []
    extra_index = []

    for i in range(len(parameter)):
        count = 0
        for ab in antibodies:
            if ab in parameter[i]:
                count += 1
            if count > 1:
                extra_suche.append(parameter[i])
                extra_index.append(i)
    for i in range(len(extra_suche)):
        extra_suche[i] = extra_suche[i].split('Ig')
        if 'beta-2-Glykoprotein-I' in extra_suche[i][0]:
            extra_suche[i][0] = extra_suche[i][0].replace('beta-2-Glykoprotein-I', 'beta-2-Glycoprotein')
        extra_suche[i] = [extra_suche[i][0] + 'Ig' + extra_suche[i][1][0], 
                          extra_suche[i][0] + 'Ig' + extra_suche[i][-1][0]]
    extra_preis = np.empty((len(extra_suche),0)).tolist()
    # Suche nach Messpreis
    for i in range(len(extra_suche)):
        for j in range(len(extra_suche[i])):
            for l in range(len(preise)):
                if extra_suche[i][j].lower() in preise.iloc[l,0].lower():
                    extra_p = preise.iloc[l,4]
                    extra_preis[i].append(extra_p)
                    break
    for i in range(len(extra_suche)):
        if len(extra_suche[i]) == len(extra_preis[i]):
            preise_dict[parameter[extra_index[i]]] = sum(extra_preis[i])
            
    # Klein-/Großschreibung
    for i in range(len(parameter)):
        if parameter[i] not in preise_dict:
            for j in range(len(preise)):
                if parameter[i].lower() in preise.iloc[j,0].lower():
                    index = preise.iloc[j,0].lower().index(parameter[i].lower())
                    para_new = preise.iloc[j,0][index: index+len(parameter[i])]
                    # preise_dict ggf. ergänzen
                    if para_new not in preise_dict:
                        add_price(list(parameter[i]))
                    parameter_new = list(map(lambda x: x.replace(parameter[i], para_new), parameter_new))
                    df['Parameter'] = parameter_new
    
    df = df.groupby(['Abteilung', 'Proben ID', 'Operator', 'Description', 'Customer', 'Parameter']).sum()
    df = df.reset_index().set_index(['Abteilung', 'Proben ID', 'Operator', 'Description', 'Customer'])
    df.drop('Gerät', axis=1, inplace=True)
    
    df['Preis/Messung [€]'] = df['Parameter'].map(preise_dict)
    df['Anzahl Messungen/Parameter'] = df['Anzahl']
    df.drop('Anzahl', axis=1, inplace=True)
    df['Anzahl Messungen*Preis [€]'] = df['Preis/Messung [€]']*df['Anzahl Messungen/Parameter']   

    writer = pd.ExcelWriter('Synlab-Rechnungen Ergebnis.xlsx', engine='xlsxwriter')
    df.to_excel(writer, sheet_name = 'Übersicht')
    writer.close()    

    # Excel-Datei Auswertung von Laborrechnungen ausfüllen

    # DM
    try:
        # Parameternamen anpassen
        para_neu = {'anti-TPO': 'a-TPO', 'anti-Tg': 'a-Tg', 'Infektionsserologie Profil: HIV-AK, HBsAG, Anti-HCV':'HIV-1/2-AK/HCV-AK/HBsAg',
                    'HBs-Ag':'HIV-1/2-AK/HCV-AK/HBsAg',
                   'fT3':'fT3/fT4', 'fT4':'fT3/fT4', 'Ana':'ANA', 'AMA Anti M2': 'AMA-Anti-M2',
                   'TRAK':'a-TSH-Rezeptor (TRAK)', }

        df_dm = df.loc['Datenmanagement/Probenmanagement',:].reset_index()
        df_dm = df_dm.groupby(['Parameter','Preis/Messung [€]'])['Anzahl Messungen/Parameter'].count()
        df_dm = df_dm.reset_index()

        para_df = list(df_dm['Parameter'])
        for i in range(len(para_df)):
            if para_df[i] in para_neu.keys():
                para_df[i] = para_neu[para_df[i]]

        df_dm['Parameter'] = para_df

        df_dm['Parameter'] = para_df
        df_dm = df_dm.groupby(['Parameter', 'Preis/Messung [€]'])['Anzahl Messungen/Parameter'].sum()
        df_dm = pd.DataFrame(df_dm).reset_index()    

        #auswertung = pd.read_excel('Auswertung von Laborrechnungen.xlsx')    

        auswertung_dm = auswertung.iloc[5:,13:16]
        auswertung_dm.columns = ['Bezeichnung', 'Preis [€]', 'Anzahl']
        auswertung_dm = auswertung_dm.reset_index().drop('index', axis=1)

        # Länge von Tabelle in Excel
        len_auswertung_dm = len(auswertung_dm) 

        for i in range(len(df_dm)):
            for ii in range(len(auswertung_dm)):
                if df_dm.iloc[i,0] == auswertung_dm.iloc[ii,0]:
                    auswertung_dm.iloc[ii,2]=df_dm.iloc[i,2]
                    auswertung_dm.iloc[ii,1]=df_dm.iloc[i,1]
        auswertung_dm = auswertung_dm.dropna().reset_index().drop('index', axis=1)
        auswertung_dm.columns = ['Parameter', 'Preis/Messung [€]', 'Anzahl Messungen/Parameter']

        para_dm = list(auswertung_dm['Parameter'])
        para_df_dm = list(df_dm['Parameter'])
        add_para = [p for p in para_df_dm if p not in para_dm]

        for p in add_para:
            df_dm_sub = df_dm[df_dm['Parameter']==p]
            auswertung_dm = pd.concat([auswertung_dm, df_dm_sub], ignore_index=True)

        # leere Zeilen in auswertung_dm einfügen, damit diese in Excel kopiert werden
        empty_rows = len_auswertung_dm-len(auswertung_dm)
        empty_df = pd.DataFrame()
        empty_df['Parameter'] = [' ']
        empty_df['Preis/Messung [€]'] = ['']
        empty_df['Anzahl Messungen/Parameter'] = ['']
        for i in range(empty_rows):
            auswertung_dm = pd.concat([auswertung_dm, empty_df], ignore_index=True)
    except:
        pass    

    # Auswertung-Tabelle ausfüllen für F&E
    try:
        df_fe = df.loc['Forschung & Entwicklung',:].reset_index()

        for i in range(len(df_fe)):
            if df_fe.iloc[i,2] == '-':
                df_fe.iloc[i,2] = df_fe.iloc[i,4]
            if df_fe.iloc[i,1] != '':
                df_fe.iloc[i,2] = df_fe.iloc[i,2] + ', ' + df_fe.iloc[i,1]
            
        df_fe = df_fe.groupby(['Description','Preis/Messung [€]'])['Anzahl Messungen/Parameter'].sum()
        df_fe = df_fe.reset_index()
        df_fe['Summe'] = df_fe['Preis/Messung [€]']*df_fe['Anzahl Messungen/Parameter']
        
        # Unterscheiden, wenn mehrere Parameter oder nur einer für Projekt vermessen wurde:
        count = df_fe['Description'].value_counts()
        count = pd.DataFrame(count).reset_index()
        one_meas = []
        for i in range(len(count)):
            if count.iloc[i,1] == 1:
                one_meas.append(count.iloc[i,0])

        df_fe_one = df_fe[df_fe['Description'].isin(one_meas)].reset_index().drop('index', axis=1).drop('Summe', axis=1)
        df_fe_one.columns = ['Bezeichnung', 'Preis', 'Anzahl']

        df_fe_more = df_fe[~df_fe['Description'].isin(one_meas)].reset_index().drop('index', axis=1)
        df_fe_more = pd.DataFrame(df_fe_more.groupby('Description').sum()['Summe']).reset_index()
        df_fe_more['Anzahl'] = [1]*len(df_fe_more)
        df_fe_more.columns = ['Bezeichnung', 'Preis', 'Anzahl']
        
        df_fe = pd.concat([df_fe_one, df_fe_more])
    except:
        pass
    
    try:
        # Produktion
        df_p = df.loc['Produktion',:].reset_index()

        dict_new_desc = dict(zip(list(df_p['Proben ID']), list(df_p['Description'])))
        dict_new_desc2 = list(dict_new_desc.values())
        dict_new_desc2 = list(set([el for el in dict_new_desc2 if el != '-']))
        dict_new_desc3 = np.empty((len(dict_new_desc2),0)).tolist()

        liste_values = list(dict_new_desc.values())
        liste_keys = list(dict_new_desc.keys())
        for i in range(len(dict_new_desc2)):
            for j in range(len(liste_values)):
                if dict_new_desc2[i] == liste_values[j]:
                    dict_new_desc3[i].append(liste_keys[j])

        for i in range(len(dict_new_desc3)):
            dict_new_desc2[i] = dict_new_desc2[i] + ' ' + str(dict_new_desc3[i])
            dict_new_desc2[i] = dict_new_desc2[i].replace("'", "").replace('[', '(').replace(']', ')')

        for i in range(len(df_p)):
            for d in dict_new_desc2:
                if df_p.iloc[i,0] in d:
                    df_p.iloc[i,2] = d
            if df_p.iloc[i,2] == '-':
                df_p.iloc[i,2] = df_p.iloc[i,4]
                    
        # Operator ergänzen
        for i in range(len(df_p)):
            if df_p.iloc[i,1] != '':
                df_p.iloc[i,2] = df_p.iloc[i,2] + ', ' + df_p.iloc[i,1]
                
        df_p_vgl = pd.DataFrame(df_p.groupby('Description').sum())[['Anzahl Messungen/Parameter', 'Preis/Messung [€]']].reset_index()
        df_p = pd.DataFrame(df_p.groupby('Description').sum()['Anzahl Messungen*Preis [€]']).reset_index()
        df_p['Anzahl'] = [1]*len(df_p)
        
        # Bei Gruppierung nach Parameter Spalten bearbeiten
        for i in range(len(df_p)):
            if df_p.iloc[i,0] == df_p_vgl.iloc[i,0]:
                if df_p_vgl.iloc[i,2]/df_p_vgl.iloc[i,1] == preise_dict[df_p.iloc[i,0][:-5]]:
                    df_p.iloc[i,1] = preise_dict[df_p.iloc[i,0][:-5]]
                    df_p.iloc[i,2] = df_p_vgl.iloc[i,1]
        
    except:
        pass
    
    # für QS
    try:
        df_qs = df.loc['Qualitätssicherung',:].reset_index()

        dict_new_desc = dict(zip(list(df_qs['Proben ID']), list(df_qs['Description'])))
        dict_new_desc2 = list(dict_new_desc.values())
        dict_new_desc2 = list(set([el for el in dict_new_desc2 if el != '-']))
        dict_new_desc3 = np.empty((len(dict_new_desc2),0)).tolist()

        liste_values = list(dict_new_desc.values())
        liste_keys = list(dict_new_desc.keys())
        for i in range(len(dict_new_desc2)):
            for j in range(len(liste_values)):
                if dict_new_desc2[i] == liste_values[j]:
                    dict_new_desc3[i].append(liste_keys[j])

        for i in range(len(dict_new_desc3)):
            dict_new_desc2[i] = dict_new_desc2[i] + ' ' + str(dict_new_desc3[i])
            dict_new_desc2[i] = dict_new_desc2[i].replace("'", "").replace('[', '(').replace(']', ')')

        for i in range(len(df_qs)):
            for d in dict_new_desc2:
                if df_qs.iloc[i,0] in d:
                    df_qs.iloc[i,2] = d
            if df_qs.iloc[i,2] == '-':
                df_qs.iloc[i,2] = df_qs.iloc[i,0]
                    
        # Operator ergänzen
        for i in range(len(df_qs)):
            if df_qs.iloc[i,1] != '':
                df_qs.iloc[i,2] = df_qs.iloc[i,2] + ', ' + df_qs.iloc[i,1]

        df_qs = pd.DataFrame(df_qs.groupby('Description').sum()['Anzahl Messungen*Preis [€]']).reset_index()
        df_qs['Anzahl'] = [1]*len(df_qs)
    except:
        pass    
    

    # insert to excel
    excel = auswertung_pfad
    workbook = openpyxl.load_workbook(excel)
    worksheet = workbook['SGS Analytics (Synlab)']
    
    def insert_to_auswertung(df, start_row, start_col):
        """ füllt die Auswertung-Datei für entsprechende Abteilung aus
        """
        num_rows, num_cols = df.shape
        data = df.values.tolist()

        for r_idx, row in enumerate(data):
            for c_idx, cell_value in enumerate(row):
                worksheet.cell(row=start_row + r_idx, column=start_col + c_idx, value=cell_value)        
            
    try:
        insert_to_auswertung(auswertung_dm, 7, 14)
    except:
        pass
    try:
        insert_to_auswertung(df_fe, 7, 20)
    except:
        pass        
    try:
        insert_to_auswertung(df_p, 7, 8)
    except:
        pass 
    try:
        insert_to_auswertung(df_qs, 7, 2)
    except:
        pass     
        
    workbook.save(excel)

    
    tada = Label(root2, text = '\nFertig! #TeamDodo kriegt von der QS einen Keks',
                 font=('Ink free',10,'bold'), bg = '#eeeee4', fg = '#869287').pack()    
    zurück = Button(root2, text = 'Zurück', bg = '#869287', font=('Ink free',12,'bold'), command = goback).pack()


    
# Auswertung-Datei auswählen und Pfad speichern
def aus_path():
    global auswertung_pfad, auswertung, dropdown_user
            
    try:
        fertig_ein.grid_forget()
    except:
        pass
    try:
        fertig_excel.grid_forget()
    except:
        pass
    
    auswertung_pfad = filedialog.askopenfilename(title = "Datei auswählen")
    auswertung = pd.read_excel(auswertung_pfad)
    
    fertig = Label(root2, text = '\nDanke, nun wähle bei Labor 28 Rechnungen zuerst noch die Abteilung \n und dann das entsprechende Labor aus.\n', bg = '#eeeee4', font=('Ink free',13))
    fertig.pack()
    options = ['F&E', 'QS', 'DM', 'P']
    selected_option = StringVar()
    dropdown_user = ttk.Combobox(root2, textvariable=selected_option, values=options, width=3)
    dropdown_user.pack()
    
    leer = Label(root2, text = '')
    leer.pack()
    
    rechnungen = Button(root2, text = 'Labor 28', font=('Ink free',12, 'bold'), width=10,
                        command = rechnungen_func,bg ='#869287')
    rechnungen.pack()

    synlab_rechn = Button(root2, text = 'Synlab', font=('Ink free',12, 'bold'), width=10, command = synlab,bg ='#869287')
    synlab_rechn.pack()    
    
# Browse-Button
def check():    
    global root2
    labor.grid_forget()
    rechnungen_check.grid_forget()

    root2 = Frame(root, width=700, height=200, bg = '#eeeee4')
    root2.grid(row=2, column=1)
    text_title = Label(root2, text='Überprüfen der Laborrechnungen', bg = '#eeeee4', font=('Ink free',14,'bold'))
    text_title.pack()
    text_info = Label(root2, text='Wähle zuerst die Auswertung-Datei zum Ausfüllen aus.\n', bg = '#eeeee4', font=('Ink free',13))
    text_info.pack()
    
    auswertung_b = Button(root2, text = 'Browse', font=('Ink free',14, 'bold'), command = aus_path,bg ='#869287')
    auswertung_b.pack()  
    
# User Interface:

# Basis
root = Tk()
root.title('Hallo QS')
root.geometry('700x730')
root.iconbitmap('O:\Forschung & Entwicklung\Allgemein\Vorlagen\Abbildungen\Dodo\dodo_icon.ico')
root.config(bg='#eeeee4')

# alle Buttons auf Startseite
hi = Label(root, text = '\nWelche Aufgabe soll der Dodo erledigen?\n',bg = '#eeeee4',font=('Ink free',20))
hi.grid(row=1, column=1)

labor = Button(root, text='Labor 28\nEinträge', width=15, font=('Ink free',12, 'bold'), command = run_eintrag,bg = '#869287')
labor.grid(row=5, column=1)

space = Label(root, text='', bg = '#eeeee4')
space.grid(row=6, column=1)

rechnungen_check = Button(root, text='Rechnungen\nüberprüfen', width=15, 
                          font=('Ink free',12, 'bold'), bg = '#869287', command=check)
rechnungen_check.grid(row=7, column=1)




# Logo
from PIL import ImageTk, Image
frame = Frame(root, width=10, height=10)
frame.grid(row=1, column=0)

img = ImageTk.PhotoImage(Image.open("O:\Forschung & Entwicklung\Allgemein\Vorlagen\Abbildungen\Dodo\dodo-dancing_ohne Hintergrund_ohne Schatten.png").resize((100,100)), master = root)
label = Label(frame, image = img, bg = '#eeeee4')
label.pack()


root.mainloop()


# In[ ]:




